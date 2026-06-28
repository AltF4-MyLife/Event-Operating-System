import base64
from io import BytesIO

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Avg, Count, Q, Sum
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import (
    AIInsight,
    AuditLog,
    Badge,
    BackupJob,
    Broadcast,
    Certificate,
    CertificateTemplate,
    CheckInLog,
    EmailLog,
    Event,
    EventModule,
    Expense,
    IncidentViolation,
    ModuleCode,
    NotificationLog,
    Participant,
    ParticipantBadge,
    PointTransaction,
    RegistrationField,
    RegistrationForm,
    RegistrationStatus,
    Refund,
    ReportSnapshot,
    SupportTicket,
    SurveyResponse,
    Ticket,
    Workshop,
    WorkshopRegistration,
    code,
    ticket_qr_code,
)


def request_meta(request):
    if not request:
        return "", ""
    return (
        request.META.get("HTTP_X_FORWARDED_FOR", request.META.get("REMOTE_ADDR", "")).split(",")[0],
        request.META.get("HTTP_USER_AGENT", "")[:1000],
    )


def qr_data_uri(value):
    try:
        import qrcode
        from qrcode.image.svg import SvgPathImage

        image = qrcode.make(value, image_factory=SvgPathImage, box_size=8)
        out = BytesIO()
        image.save(out)
        return "data:image/svg+xml;base64," + base64.b64encode(out.getvalue()).decode("ascii")
    except Exception:
        return ""


def plain_json(value):
    if isinstance(value, dict):
        return {key: plain_json(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [plain_json(item) for item in value]
    if hasattr(value, "pk"):
        return str(value.pk)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def audit(action, obj=None, actor=None, before=None, after=None, request=None, note=""):
    ip, user_agent = request_meta(request)
    event = getattr(obj, "event", None)
    organization = getattr(obj, "organization", None) or getattr(event, "organization", None)
    return AuditLog.objects.create(
        actor=actor if getattr(actor, "is_authenticated", False) else None,
        action=action,
        entity_type=obj.__class__.__name__ if obj else "System",
        entity_id=str(getattr(obj, "pk", "")) if obj else "",
        organization=organization,
        event=event,
        ip_address=ip or None,
        user_agent=user_agent,
        before=plain_json(before or {}),
        after=plain_json(after or {}),
        note=note,
    )


def ensure_default_modules(event, actor=None):
    created_or_updated = []
    for module_code, label in ModuleCode.choices:
        module, _ = EventModule.objects.get_or_create(
            event=event,
            code=module_code,
            defaults={"label": label, "enabled": True, "changed_by": actor},
        )
        if not module.label:
            module.label = label
            module.save(update_fields=["label", "updated_at"])
        created_or_updated.append(module)
    return created_or_updated


def ensure_registration_form(event):
    form, created = RegistrationForm.objects.get_or_create(
        event=event,
        defaults={
            "title": f"نموذج تسجيل {event.name}",
            "description": "املأ البيانات المطلوبة لإتمام التسجيل.",
            "duplicate_rules": {"phone": True, "email": True, "national_id": True},
        },
    )
    if created:
        default_fields = [
            ("الاسم الكامل", "full_name", RegistrationField.FieldType.TEXT, True, 1),
            ("رقم الهاتف", "phone", RegistrationField.FieldType.PHONE, True, 2),
            ("البريد الإلكتروني", "email", RegistrationField.FieldType.EMAIL, False, 3),
            ("المدرسة", "school_name", RegistrationField.FieldType.TEXT, False, 4),
            ("الإدارة التعليمية", "administration_name", RegistrationField.FieldType.TEXT, False, 5),
            ("المحافظة", "governorate", RegistrationField.FieldType.TEXT, False, 6),
            ("سبب الحضور", "reason", RegistrationField.FieldType.TEXTAREA, False, 7),
            ("الموافقة على الشروط", "terms", RegistrationField.FieldType.CONSENT, True, 8),
        ]
        for label, key, field_type, required, order in default_fields:
            RegistrationField.objects.create(
                form=form,
                label=label,
                key=key,
                field_type=field_type,
                required=required,
                order=order,
            )
    return form


def participant_duplicate_hash(event_id, phone="", email="", national_id=""):
    import hashlib

    base = f"{event_id}:{phone.lower()}:{email.lower()}:{national_id.lower()}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def validate_dynamic_fields(form, answers):
    missing = []
    for field in form.fields.filter(is_active=True, required=True):
        value = answers.get(field.key)
        if field.field_type == RegistrationField.FieldType.CONSENT:
            valid = value in {True, "true", "on", "1", "yes", "موافق"}
        else:
            valid = value not in {None, ""}
        if not valid:
            missing.append(field.label)
    if missing:
        raise ValidationError(f"حقول مطلوبة ناقصة: {', '.join(missing)}")


@transaction.atomic
def register_participant(event, data, request=None):
    ensure_default_modules(event)
    form = ensure_registration_form(event)
    if event.maintenance_mode or not event.module_enabled(ModuleCode.PUBLIC_REGISTRATION):
        raise ValidationError("التسجيل مغلق مؤقتًا لهذه الفعالية.")
    if not event.registration_open or not form.is_open:
        raise ValidationError("نموذج التسجيل مغلق حاليًا.")
    if event.capacity and event.participants.filter(is_deleted=False).count() >= event.capacity:
        if event.module_enabled(ModuleCode.WAITLIST):
            initial_status = RegistrationStatus.WAITLISTED
        else:
            raise ValidationError("اكتملت سعة الفعالية.")
    elif event.module_enabled(ModuleCode.AUTO_APPROVAL):
        initial_status = RegistrationStatus.APPROVED
    elif event.module_enabled(ModuleCode.MANUAL_REVIEW):
        initial_status = RegistrationStatus.PENDING_REVIEW
    else:
        initial_status = RegistrationStatus.SUBMITTED

    answers = data.get("dynamic_answers") or data.get("answers") or {}
    validate_dynamic_fields(form, {**data, **answers})
    duplicate_hash = participant_duplicate_hash(
        event.id,
        data.get("phone", ""),
        data.get("email", ""),
        data.get("national_id", ""),
    )
    if Participant.objects.filter(event=event, duplicate_hash=duplicate_hash, is_deleted=False).exists():
        raise ValidationError("يوجد تسجيل سابق بنفس بيانات التواصل.")

    participant = Participant.objects.create(
        event=event,
        registration_type=data.get("registration_type", Participant.RegistrationType.STUDENT),
        full_name=data["full_name"].strip(),
        phone=data["phone"].strip(),
        email=data.get("email", "").strip(),
        governorate=data.get("governorate", "").strip(),
        national_id=data.get("national_id", "").strip(),
        age=data.get("age") or None,
        gender=data.get("gender", "").strip(),
        guardian_name=data.get("guardian_name", "").strip(),
        guardian_phone=data.get("guardian_phone", "").strip(),
        reason=data.get("reason", "").strip(),
        status=initial_status,
        dynamic_answers=answers,
    )
    audit("participant.registered", participant, request=request, after={"status": participant.status})
    return participant


@transaction.atomic
def review_participant(participant, action, actor=None, note="", request=None):
    transitions = {
        "approve": RegistrationStatus.APPROVED,
        "reject": RegistrationStatus.REJECTED,
        "waitlist": RegistrationStatus.WAITLISTED,
        "need_info": RegistrationStatus.NEED_MORE_INFO,
        "ban": RegistrationStatus.BANNED,
        "block_certificate": RegistrationStatus.CERTIFICATE_BLOCKED,
    }
    if action not in transitions:
        raise ValidationError("إجراء مراجعة غير معروف.")
    before = {"status": participant.status}
    participant.status = transitions[action]
    if action == "approve":
        participant.approved_at = timezone.now()
    if action == "block_certificate":
        participant.certificate_blocked = True
        participant.certificate_block_reason = note or "تم منع الشهادة من الإدارة."
    if action == "ban":
        participant.is_blacklisted = True
    participant.review_notes = note or participant.review_notes
    participant.save()
    audit("participant.reviewed", participant, actor=actor, before=before, after={"status": participant.status}, request=request, note=note)
    return participant


@transaction.atomic
def bulk_review_participants(participants, action, actor=None, note="", request=None):
    count = 0
    for participant in participants.select_for_update():
        review_participant(participant, action, actor=actor, note=note, request=request)
        count += 1
    return count


def normalized_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def pick_value(row, aliases, default=""):
    for alias in aliases:
        if alias in row and row[alias] not in {None, ""}:
            return str(row[alias]).strip()
    return default


@transaction.atomic
def import_participants_workbook(event, file_obj, actor=None, request=None, default_status=RegistrationStatus.SUBMITTED):
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValidationError("ملف الاستيراد فارغ.")
    headers = [normalized_header(value) for value in rows[0]]
    created = 0
    skipped = 0
    errors = []
    for index, values in enumerate(rows[1:], start=2):
        row = {headers[pos]: values[pos] for pos in range(min(len(headers), len(values)))}
        full_name = pick_value(row, ["full_name", "name", "الاسم", "الاسم_الكامل"])
        phone = pick_value(row, ["phone", "mobile", "رقم_الهاتف", "الهاتف"])
        email = pick_value(row, ["email", "البريد", "البريد_الإلكتروني"])
        national_id = pick_value(row, ["national_id", "الرقم_القومي"])
        governorate = pick_value(row, ["governorate", "المحافظة"])
        registration_type = pick_value(row, ["registration_type", "type", "نوع_التسجيل"], Participant.RegistrationType.STUDENT)
        if registration_type not in Participant.RegistrationType.values:
            registration_type = Participant.RegistrationType.STUDENT
        if not full_name or not phone:
            errors.append({"row": index, "error": "الاسم ورقم الهاتف مطلوبان."})
            skipped += 1
            continue
        duplicate_hash = participant_duplicate_hash(event.id, phone, email, national_id)
        if Participant.objects.filter(event=event, duplicate_hash=duplicate_hash, is_deleted=False).exists():
            errors.append({"row": index, "error": "تسجيل مكرر."})
            skipped += 1
            continue
        Participant.objects.create(
            event=event,
            full_name=full_name,
            phone=phone,
            email=email,
            national_id=national_id,
            governorate=governorate,
            registration_type=registration_type,
            status=default_status,
            dynamic_answers={"imported": True, "source_row": index},
        )
        created += 1
    audit("participants.imported", event, actor=actor, request=request, after={"created": created, "skipped": skipped, "errors": errors[:20]})
    return {"created": created, "skipped": skipped, "errors": errors}


@transaction.atomic
def perform_checkin(event, scanned_code, action=CheckInLog.Action.CHECKIN, actor=None, gate="", device="", workshop=None, request=None):
    ensure_default_modules(event)
    if action in {CheckInLog.Action.CHECKIN, CheckInLog.Action.CHECKOUT} and not event.module_enabled(ModuleCode.CHECKIN):
        raise ValidationError("خاصية تسجيل الحضور غير مفعلة.")
    if action == CheckInLog.Action.WORKSHOP and not event.module_enabled(ModuleCode.WORKSHOP_ATTENDANCE):
        raise ValidationError("خاصية حضور الورش غير مفعلة.")
    participant = Participant.objects.filter(
        event=event,
        is_deleted=False,
    ).filter(Q(qr_code=scanned_code) | Q(tracking_code=scanned_code)).first()
    if not participant:
        log = CheckInLog.objects.create(event=event, action=action, code_scanned=scanned_code, gate=gate, device=device, performed_by=actor, success=False, message="QR غير صالح")
        audit("checkin.failed", log, actor=actor, request=request, note=log.message)
        raise ValidationError("QR أو كود المتابعة غير صالح.")
    if participant.is_blacklisted or participant.status == RegistrationStatus.BANNED:
        log = CheckInLog.objects.create(event=event, participant=participant, action=action, code_scanned=scanned_code, gate=gate, device=device, performed_by=actor, success=False, message="المشارك محظور")
        audit("checkin.blocked", log, actor=actor, request=request, note=log.message)
        raise ValidationError("لا يمكن تسجيل حضور مشارك محظور.")
    if participant.status not in {RegistrationStatus.APPROVED, RegistrationStatus.CHECKED_IN, RegistrationStatus.CHECKED_OUT, RegistrationStatus.ATTENDED}:
        log = CheckInLog.objects.create(event=event, participant=participant, action=action, code_scanned=scanned_code, gate=gate, device=device, performed_by=actor, success=False, message="المشارك غير مقبول")
        audit("checkin.failed", log, actor=actor, request=request, note=log.message)
        raise ValidationError("المشارك غير مقبول للحضور.")
    if action == CheckInLog.Action.CHECKIN and event.prevent_duplicate_checkin and participant.checked_in_at:
        log = CheckInLog.objects.create(event=event, participant=participant, action=action, code_scanned=scanned_code, gate=gate, device=device, performed_by=actor, success=False, message="حضور مكرر")
        audit("checkin.duplicate", log, actor=actor, request=request, note=log.message)
        raise ValidationError("تم تسجيل حضور هذا المشارك من قبل.")
    if action == CheckInLog.Action.WORKSHOP:
        if not workshop:
            raise ValidationError("يجب تحديد الورشة.")
        registration = WorkshopRegistration.objects.filter(workshop=workshop, participant=participant).first()
        if not registration or registration.status not in {WorkshopRegistration.Status.REGISTERED, WorkshopRegistration.Status.ATTENDED}:
            raise ValidationError("المشارك غير مسجل في هذه الورشة.")
        registration.status = WorkshopRegistration.Status.ATTENDED
        registration.attended_at = timezone.now()
        registration.save()
        if workshop.points:
            award_points(participant, workshop.points, f"حضور ورشة {workshop.title}", actor=actor, source="workshop")
    elif action == CheckInLog.Action.CHECKOUT:
        participant.status = RegistrationStatus.CHECKED_OUT
        participant.checked_out_at = timezone.now()
        participant.save(update_fields=["status", "checked_out_at", "updated_at"])
    else:
        participant.status = RegistrationStatus.CHECKED_IN
        participant.checked_in_at = timezone.now()
        participant.save(update_fields=["status", "checked_in_at", "updated_at"])
        checkin_rule = participant.event.point_rules.filter(trigger="checkin", enabled=True).first()
        if checkin_rule:
            award_points(participant, checkin_rule.value, checkin_rule.name, actor=actor, source="checkin")

    log = CheckInLog.objects.create(
        event=event,
        participant=participant,
        workshop=workshop,
        action=action,
        code_scanned=scanned_code,
        gate=gate,
        device=device,
        performed_by=actor,
        success=True,
        message="تم التسجيل بنجاح",
    )
    audit("checkin.success", log, actor=actor, request=request)
    return log


@transaction.atomic
def enroll_workshop(workshop, participant, actor=None, request=None):
    if participant.event_id != workshop.event_id:
        raise ValidationError("المشارك لا يتبع نفس الفعالية.")
    if not workshop.registration_open:
        raise ValidationError("التسجيل في الورشة مغلق.")
    status = WorkshopRegistration.Status.REGISTERED
    if workshop.capacity and workshop.seats_available <= 0:
        if not workshop.waitlist_enabled:
            raise ValidationError("اكتملت سعة الورشة.")
        status = WorkshopRegistration.Status.WAITLISTED
    registration, created = WorkshopRegistration.objects.get_or_create(
        workshop=workshop,
        participant=participant,
        defaults={"status": status},
    )
    if not created and registration.status == WorkshopRegistration.Status.CANCELLED:
        registration.status = status
        registration.save()
    audit("workshop.enrolled", registration, actor=actor, request=request, after={"status": registration.status})
    return registration


def participant_points_total(participant):
    return participant.point_transactions.aggregate(total=Sum("value")).get("total") or 0


def award_points(participant, value, reason, actor=None, source="manual"):
    tx = PointTransaction.objects.create(participant=participant, value=value, reason=reason, awarded_by=actor, source=source)
    apply_auto_badges(participant, actor=actor)
    return tx


def apply_auto_badges(participant, actor=None):
    total = participant_points_total(participant)
    for badge in Badge.objects.filter(event=participant.event, auto_award=True):
        if badge.points_required and total < badge.points_required:
            continue
        ParticipantBadge.objects.get_or_create(participant=participant, badge=badge, defaults={"awarded_by": actor, "reason": "منح تلقائي"})


@transaction.atomic
def record_violation(participant, violation_type, reported_by=None, notes="", location=""):
    violation = IncidentViolation.objects.create(
        event=participant.event,
        participant=participant,
        violation_type=violation_type,
        reported_by=reported_by,
        notes=notes,
        location=location,
        action_taken=violation_type.default_action if violation_type else "",
    )
    if violation_type and violation_type.points_penalty:
        award_points(participant, -abs(violation_type.points_penalty), f"مخالفة: {violation_type.name}", actor=reported_by, source="violation")
    if violation_type and violation_type.blocks_certificate:
        participant.certificate_blocked = True
        participant.certificate_block_reason = f"مخالفة: {violation_type.name}"
        participant.status = RegistrationStatus.CERTIFICATE_BLOCKED
        participant.save(update_fields=["certificate_blocked", "certificate_block_reason", "status", "updated_at"])
    audit("violation.recorded", violation, actor=reported_by, note=notes)
    return violation


def attendance_percent(participant):
    event_days = max(participant.event.days.count(), 1)
    attended = participant.checkin_logs.filter(action=CheckInLog.Action.CHECKIN, success=True).dates("checked_at", "day").count()
    return int((attended / event_days) * 100)


def certificate_eligibility(participant):
    event = participant.event
    if not event.module_enabled(ModuleCode.CERTIFICATES):
        return False, "الشهادات غير مفعلة لهذه الفعالية."
    if participant.certificate_blocked:
        return False, participant.certificate_block_reason or "الشهادة محظورة."
    if participant.violations.filter(status__in=[IncidentViolation.Status.OPEN, IncidentViolation.Status.APPEALED], violation_type__blocks_certificate=True).exists():
        return False, "يوجد مخالفة تمنع إصدار الشهادة."
    if participant.status not in {RegistrationStatus.CHECKED_IN, RegistrationStatus.CHECKED_OUT, RegistrationStatus.ATTENDED, RegistrationStatus.APPROVED}:
        return False, "المشارك غير مؤهل حسب حالة التسجيل."
    if participant.checked_in_at and attendance_percent(participant) < event.min_attendance_percent_for_certificate:
        return False, "نسبة الحضور أقل من الحد المطلوب."
    feedback_required = event.require_feedback_for_certificate or event.surveys.filter(active=True, certificate_required=True).exists()
    if feedback_required and not SurveyResponse.objects.filter(survey__event=event, participant=participant).exists():
        return False, "يجب إكمال التقييم أولًا."
    return True, "مؤهل لإصدار الشهادة."


@transaction.atomic
def issue_certificate(participant, actor=None, template=None, certificate_type="attendance"):
    eligible, reason = certificate_eligibility(participant)
    if not eligible:
        raise ValidationError(reason)
    template = template or CertificateTemplate.objects.filter(event=participant.event, certificate_type=certificate_type, active=True).first()
    if not template:
        template = CertificateTemplate.objects.create(event=participant.event, name="الشهادة الافتراضية", certificate_type=certificate_type)
    certificate, _ = Certificate.objects.get_or_create(
        event=participant.event,
        participant=participant,
        certificate_type=certificate_type,
        defaults={"template": template},
    )
    certificate.status = Certificate.Status.ISSUED
    certificate.template = template
    certificate.issued_at = timezone.now()
    certificate.save()
    audit("certificate.issued", certificate, actor=actor)
    return certificate


def render_certificate_pdf(certificate):
    html = render_to_string("core/pdf_certificate.html", {"certificate": certificate, "qr_image": qr_data_uri(certificate.verification_code)})
    try:
        from weasyprint import HTML

        return HTML(string=html, base_url=str(settings.BASE_DIR)).write_pdf()
    except Exception:
        # Local Windows machines can miss native rendering pieces. Keep the action auditable and non-crashing.
        return html.encode("utf-8")


def send_event_email(event, to_email, subject, body, participant=None):
    provider = getattr(settings, "EMAIL_PROVIDER", "mock")
    log = EmailLog.objects.create(event=event, participant=participant, provider=provider, to_email=to_email, subject=subject, body=body, attempts=1)
    provider_ready = (provider == "brevo" and settings.BREVO_API_KEY) or (provider == "sendgrid" and settings.SENDGRID_API_KEY)
    if not provider_ready and provider != "smtp":
        log.status = EmailLog.Status.MOCKED
        log.error = "No external provider configured; message recorded as a safe mock."
        log.save(update_fields=["status", "error", "updated_at"])
        return log
    try:
        send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [to_email], fail_silently=False)
        log.status = EmailLog.Status.SENT
    except Exception as exc:
        log.status = EmailLog.Status.FAILED
        log.error = str(exc)
    log.save(update_fields=["status", "error", "updated_at"])
    return log


class SafeTemplateContext(dict):
    def __missing__(self, key):
        return ""


def render_message_template(text, event, participant=None, extra=None):
    context = SafeTemplateContext(
        event=event.name if event else "",
        organization=event.organization.name if event and event.organization_id else "",
        participant=participant.full_name if participant else "",
        tracking_code=participant.tracking_code if participant else "",
        qr_code=participant.qr_code if participant else "",
        status=participant.get_status_display() if participant else "",
    )
    context.update(extra or {})
    try:
        return (text or "").format_map(context)
    except ValueError:
        return text or ""


def send_templated_email(template, participant, actor=None):
    if not participant.email:
        raise ValidationError("لا يوجد بريد إلكتروني للمشارك.")
    subject = render_message_template(template.subject, template.event, participant)
    body = render_message_template(template.body, template.event, participant)
    log = send_event_email(template.event, participant.email, subject, body, participant=participant)
    audit("email.template_sent", log, actor=actor)
    return log


def send_certificate_email(certificate):
    participant = certificate.participant
    if not participant.email:
        raise ValidationError("لا يوجد بريد إلكتروني للمشارك.")
    verify_path = f"/certificates/verify/{certificate.verification_code}/"
    body = (
        f"مرحبًا {participant.full_name}\n"
        f"تم إصدار شهادتك في {certificate.event.name}.\n"
        f"رقم الشهادة: {certificate.serial_number}\n"
        f"رابط التحقق: {verify_path}"
    )
    log = send_event_email(certificate.event, participant.email, f"شهادة {certificate.event.name}", body, participant=participant)
    certificate.sent_count += 1
    certificate.save(update_fields=["sent_count", "updated_at"])
    return log


def send_push(event, title, body, audience="", payload=None, recipient_user=None):
    safe_payload = payload or {}
    blocked_keys = {"national_id", "phone", "email", "guardian_phone"}
    for key in blocked_keys:
        safe_payload.pop(key, None)
    return NotificationLog.objects.create(
        event=event,
        recipient_user=recipient_user,
        audience=audience,
        title=title,
        body=body[:240],
        safe_payload=safe_payload,
        status="mocked",
    )


def send_broadcast(broadcast):
    if broadcast.channel == Broadcast.Channel.EMAIL:
        recipients = Participant.objects.filter(event=broadcast.event)
        for participant in recipients.exclude(email=""):
            send_event_email(broadcast.event, participant.email, broadcast.title, broadcast.message, participant=participant)
    elif broadcast.channel == Broadcast.Channel.PUSH:
        send_push(broadcast.event, broadcast.title, broadcast.message, audience=broadcast.audience)
    broadcast.sent_at = timezone.now()
    broadcast.status = "sent"
    broadcast.save(update_fields=["sent_at", "status", "updated_at"])
    return broadcast


def dashboard_stats(event=None, organization=None):
    events = Event.objects.all()
    if organization:
        events = events.filter(organization=organization)
    if event:
        events = events.filter(pk=event.pk)
    participants = Participant.objects.filter(event__in=events, is_deleted=False)
    checkins = CheckInLog.objects.filter(event__in=events, action=CheckInLog.Action.CHECKIN, success=True)
    total = participants.count()
    checked = checkins.values("participant_id").distinct().count()
    satisfaction = SurveyResponse.objects.filter(survey__event__in=events).aggregate(avg=Avg("satisfaction_score")).get("avg") or 0
    return {
        "organizations": organization and 1 or events.values("organization_id").distinct().count(),
        "events": events.count(),
        "participants": total,
        "approved": participants.filter(status=RegistrationStatus.APPROVED).count(),
        "rejected": participants.filter(status=RegistrationStatus.REJECTED).count(),
        "waitlisted": participants.filter(status=RegistrationStatus.WAITLISTED).count(),
        "checked_in": checked,
        "attendance_rate": round((checked / total) * 100, 1) if total else 0,
        "workshops": Workshop.objects.filter(event__in=events).count(),
        "certificates": Certificate.objects.filter(event__in=events).count(),
        "violations": IncidentViolation.objects.filter(event__in=events).count(),
        "support_tickets": SupportTicket.objects.filter(event__in=events).count(),
        "sos": events.aggregate(count=Count("sos_reports")).get("count") or 0,
        "satisfaction": round(satisfaction, 1),
    }


def export_participants_workbook(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "المشاركون"
    headers = ["الكود", "الاسم", "الهاتف", "البريد", "النوع", "الحالة", "الفعالية", "المحافظة", "تاريخ التسجيل"]
    ws.append(headers)
    for participant in queryset.select_related("event"):
        ws.append([
            participant.tracking_code,
            participant.full_name,
            participant.phone,
            participant.email,
            participant.get_registration_type_display(),
            participant.get_status_display(),
            participant.event.name,
            participant.governorate,
            participant.created_at.strftime("%Y-%m-%d %H:%M"),
        ])
    return wb


def export_attendance_workbook(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "الحضور"
    ws.append(["الفعالية", "المشارك", "الكود", "الإجراء", "نجاح", "البوابة", "الجهاز", "الموظف", "الوقت", "الرسالة"])
    for log in queryset.select_related("event", "participant", "performed_by"):
        ws.append([
            log.event.name,
            log.participant.full_name if log.participant else "",
            log.code_scanned,
            log.get_action_display(),
            "نعم" if log.success else "لا",
            log.gate,
            log.device,
            log.performed_by.username if log.performed_by else "",
            log.checked_at.strftime("%Y-%m-%d %H:%M"),
            log.message,
        ])
    return wb


def export_certificates_workbook(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "الشهادات"
    ws.append(["الفعالية", "المشارك", "النوع", "الحالة", "Serial", "Verification", "تاريخ الإصدار", "مرات الإرسال", "مرات التحميل"])
    for certificate in queryset.select_related("event", "participant"):
        ws.append([
            certificate.event.name,
            certificate.participant.full_name,
            certificate.certificate_type,
            certificate.get_status_display(),
            certificate.serial_number,
            certificate.verification_code,
            certificate.issued_at.strftime("%Y-%m-%d %H:%M"),
            certificate.sent_count,
            certificate.download_count,
        ])
    return wb


def export_workshops_workbook(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "الورش"
    ws.append(["الفعالية", "الورشة", "المدرب", "القاعة", "البداية", "النهاية", "السعة", "المسجلون", "المقاعد المتاحة", "النقاط", "التسجيل مفتوح"])
    for workshop in queryset.select_related("event", "trainer", "hall").prefetch_related("registrations"):
        ws.append([
            workshop.event.name,
            workshop.title,
            workshop.trainer.name if workshop.trainer else "",
            workshop.hall.name if workshop.hall else "",
            workshop.starts_at.strftime("%Y-%m-%d %H:%M"),
            workshop.ends_at.strftime("%Y-%m-%d %H:%M"),
            workshop.capacity,
            workshop.seats_taken,
            workshop.seats_available,
            workshop.points,
            "نعم" if workshop.registration_open else "لا",
        ])
    return wb


def export_violations_workbook(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "المخالفات"
    ws.append(["الفعالية", "المشارك", "نوع المخالفة", "المكان", "الإجراء", "الحالة", "المشرف", "ملاحظات", "التاريخ"])
    for violation in queryset.select_related("event", "participant", "violation_type", "reported_by"):
        ws.append([
            violation.event.name,
            violation.participant.full_name,
            violation.violation_type.name if violation.violation_type else "",
            violation.location,
            violation.action_taken,
            violation.get_status_display(),
            violation.reported_by.username if violation.reported_by else "",
            violation.notes,
            violation.created_at.strftime("%Y-%m-%d %H:%M"),
        ])
    return wb


def workbook_bytes(workbook):
    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    return out


def final_report_data(event):
    participants = event.participants.filter(is_deleted=False)
    tickets = Ticket.objects.filter(event=event)
    expenses = Expense.objects.filter(event=event)
    revenue = tickets.aggregate(total=Sum("amount")).get("total") or 0
    expense_total = expenses.aggregate(total=Sum("amount")).get("total") or 0
    refund_total = Refund.objects.filter(event=event, status=Refund.Status.PAID).aggregate(total=Sum("amount")).get("total") or 0
    return {
        "stats": dashboard_stats(event=event),
        "by_status": list(participants.values("status").annotate(total=Count("id")).order_by("status")),
        "by_type": list(participants.values("registration_type").annotate(total=Count("id")).order_by("registration_type")),
        "top_schools": list(participants.exclude(school=None).values("school__name").annotate(total=Count("id")).order_by("-total")[:10]),
        "workshops": list(event.workshops.annotate(registrations_total=Count("registrations")).values("title", "capacity", "registrations_total")),
        "certificates": list(event.certificates.values("status").annotate(total=Count("id")).order_by("status")),
        "feedback_avg": SurveyResponse.objects.filter(survey__event=event).aggregate(avg=Avg("satisfaction_score")).get("avg") or 0,
        "finance": {
            "revenue": float(revenue),
            "expenses": float(expense_total),
            "refunds": float(refund_total),
            "net": float(revenue - expense_total - refund_total),
            "tickets": tickets.count(),
        },
        "risks": {
            "open_sos": event.sos_reports.exclude(status__in=["closed", "false_alarm"]).count(),
            "open_support": event.support_tickets.exclude(status__in=["closed", "resolved"]).count(),
            "violations": event.violations.count(),
        },
    }


def create_report_snapshot(event, actor=None, report_type="final"):
    data = final_report_data(event)
    snapshot = ReportSnapshot.objects.create(
        event=event,
        report_type=report_type,
        title=f"تقرير {event.name}",
        data=data,
        generated_by=actor if getattr(actor, "is_authenticated", False) else None,
    )
    audit("report.snapshot_created", snapshot, actor=actor)
    return snapshot


def render_event_report_pdf(event, actor=None):
    snapshot = create_report_snapshot(event, actor=actor)
    html = render_to_string("core/pdf_event_report.html", {"event": event, "snapshot": snapshot, "data": snapshot.data})
    try:
        from weasyprint import HTML

        return HTML(string=html, base_url=str(settings.BASE_DIR)).write_pdf()
    except Exception:
        return html.encode("utf-8")


def create_backup_snapshot(organization=None, actor=None):
    events = Event.objects.filter(organization=organization) if organization else Event.objects.all()
    data = {
        "organization": organization.name if organization else "platform",
        "created_at": timezone.now().isoformat(),
        "events": events.count(),
        "participants": Participant.objects.filter(event__in=events, is_deleted=False).count(),
        "certificates": Certificate.objects.filter(event__in=events).count(),
        "tickets": Ticket.objects.filter(event__in=events).count(),
        "support_tickets": SupportTicket.objects.filter(event__in=events).count(),
    }
    job = BackupJob.objects.create(organization=organization, job_type="metadata_snapshot", status=BackupJob.Status.DONE)
    audit("backup.snapshot_created", job, actor=actor, after=data)
    return job, data


def cancel_ticket(ticket, actor=None, request=None, note=""):
    before = {"cancelled": ticket.cancelled, "payment_status": ticket.payment_status}
    ticket.cancelled = True
    ticket.payment_status = Ticket.PaymentStatus.FAILED if ticket.payment_status == Ticket.PaymentStatus.PENDING else ticket.payment_status
    ticket.save(update_fields=["cancelled", "payment_status", "updated_at"])
    audit("ticket.cancelled", ticket, actor=actor, request=request, before=before, after={"cancelled": ticket.cancelled, "payment_status": ticket.payment_status}, note=note)
    return ticket


def reissue_ticket(ticket, actor=None, request=None):
    before = {"qr_code": ticket.qr_code, "cancelled": ticket.cancelled}
    ticket.qr_code = ticket_qr_code()
    ticket.cancelled = False
    ticket.save(update_fields=["qr_code", "cancelled", "updated_at"])
    audit("ticket.reissued", ticket, actor=actor, request=request, before=before, after={"qr_code": ticket.qr_code, "cancelled": ticket.cancelled})
    return ticket


def update_ticket_payment(ticket, payment_status, actor=None, request=None):
    if payment_status not in Ticket.PaymentStatus.values:
        raise ValidationError("حالة الدفع غير معروفة.")
    before = {"payment_status": ticket.payment_status}
    ticket.payment_status = payment_status
    if payment_status == Ticket.PaymentStatus.PAID and not ticket.receipt_number:
        ticket.receipt_number = f"RCT-{ticket.qr_code.split('-')[-1]}"
    ticket.save(update_fields=["payment_status", "receipt_number", "updated_at"])
    audit("ticket.payment_updated", ticket, actor=actor, request=request, before=before, after={"payment_status": ticket.payment_status})
    return ticket


def rule_based_duplicate_insights(event):
    created = []
    duplicates = (
        Participant.objects.filter(event=event, is_deleted=False)
        .values("duplicate_hash")
        .annotate(total=Count("id"))
        .filter(total__gt=1)
    )
    for item in duplicates:
        participants = list(Participant.objects.filter(event=event, duplicate_hash=item["duplicate_hash"]).values_list("tracking_code", flat=True))
        insight, _ = AIInsight.objects.get_or_create(
            event=event,
            insight_type="duplicate_registration",
            evidence={"tracking_codes": participants},
            defaults={
                "score": min(item["total"] / 5, 1),
                "summary": f"تسجيلات محتملة التكرار: {', '.join(participants)}",
            },
        )
        created.append(insight)
    return created
